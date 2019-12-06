from openpyxl.styles import Alignment


fontObj1 = Font(name=u'Arial', bold=False, italic=False, size=8)
align = Alignment(vertical='top')
# Remove Duplication
workbook = openpyxl.load_workbook('EAP_export_of_EDPs_401.xlsx')
worksheet = workbook.worksheets[0]
worksheet.font = fontObj1
max_row = worksheet.max_row
max_col = worksheet.max_column
value = []
num_consistency = max_row - 1
for x in range(6, max_row):
    cell_data = worksheet.cell(row=x, column=2).value
    value.append(cell_data)
dup_index = []
dic = list(set(value))
b = 0
# Get the index with same procedure name
for i in range(0, len(dic)):
    b += 1
    if b%1000 == 0:
        print(b)
    a = []
    for j in range(0, len(value)):
        if value[j] == dic[i]:
            a.append(j)
            dup_index.append(a)
for i in range(0, len(dup_index)):
    for j in range(0, len(dup_index[i])):
        num = dup_index[i][j] + 6
        data = worksheet.cell(row=num, column=9).value
        if len(data[3:]) < 6 or len(data[3:]) > 18 or data[0:3] != 'SHX':
            cell_value = worksheet.cell(row=num, column=3, value='*SD')
            cell_value.font = fontObj1
            cell_value.alignment = align
# workbook.save('EAP_export_of_EDPs_401(1).xlsx')
# print(dup_index)
# In ORP not in EAP
workbook1 = openpyxl.load_workbook('EAP_export_of_EDPs_401 - InORPbutNotInEAP.xlsx')
worksheet1 = workbook1.worksheets[0]
max_row1 = worksheet1.max_row
max_col1 = worksheet1.max_column
name, name_ID = [], []
for y in range(2, max_row1):
    cell_data1 = worksheet1.cell(row=y, column=2).value
    cell_data0 = worksheet1.cell(row=y, column=1).value
    name.append(cell_data1)
    name_ID.append(cell_data0)
data_consistency = []
for column in range(15, max_col):
    data = worksheet.cell(row=max_row-1, column=column).value
    data_consistency.append(data)
b = 0
# for i in range(0, len(value)):
#     # print(value[i])
#     if value[i] in name:
#         # print(value[i])
#         b += 1
#         index_ORP = name.index(value[i])
#         num = i + 6
#         cell_value0 = worksheet.cell(row=num, column=1).value
#         worksheet.cell(row=num, column=1, value='*' + cell_value0)
#         worksheet.cell(row=num, column=5, value='Update')
#         worksheet.cell(row=num, column=9, value='SHX' + name_ID[index_ORP])
letter = ['A', 'B', 'D', 'E', 'H', 'I', 'L', 'N']
for i in range(0, len(name)):
    num = max_row + i + 1
    num_str = str(num)
    for k in range(0, len(letter)):
        worksheet[letter[k] + num_str].font = fontObj1
    worksheet.cell(row=num, column=1, value='*')
    worksheet.cell(row=num, column=2, value=name[i])
    worksheet.cell(row=num, column=4, value=name[i])
    worksheet.cell(row=num, column=9, value='SHX' + name_ID[i])
    worksheet.cell(row=num, column=5, value='Update')
    worksheet.cell(row=num, column=8, value='INACTIVE')
    worksheet.cell(row=num, column=12, value=name[i])
    worksheet.cell(row=num, column=14, value=name[i])
    for j in range(0, len(data_consistency)):
        num1 = j + 15
        o = worksheet.cell(row=num, column=num1, value=data_consistency[j])
        o.font = fontObj1
# workbook.save('EAP_export_of_EDPs_401(1).xlsx')
# In EAP not in ORP
workbook2 = openpyxl.load_workbook('EAP_export_of_EDPs_401 - InEAPbutNotInORP.xlsx')
worksheet2 = workbook2.worksheets[0]
max_row2 = worksheet2.max_row
max_col2 = worksheet2.max_column
procedure_num, procedure_correction = [], []
name_EAP = []
for x in range(2, max_row2):
    # cell_data = worksheet2.cell(row=x, column=4).value
    cell_correction = worksheet2.cell(row=x, column=5).value
    # if cell_correction == '':
    # cell_correction = worksheet2.cell(row=x, column=4).value
    EAP = worksheet2.cell(row=x, column=2).value
    # procedure_num.append(cell_data)
    procedure_correction.append(cell_correction)
    name_EAP.append(EAP)
# print(procedure_correction)
c = 0
d = 0
for i in range(0, len(value)):
    # print(value[i])
    if value[i] in name_EAP:
        index_EAP = name_EAP.index(value[i])
        # c += 1
        # print(c)
        num = i + 6
        if procedure_correction[index_EAP] is not None:
            # d += 1
            # print(d)
            o = worksheet.cell(row=num, column=9, value=procedure_correction[index_EAP])
            o.font = fontObj1
        else:
            # c += 1
            # print(c)
            o = worksheet.cell(row=num, column=3, value='*SD')
            o.font = fontObj1
            o.alignment = align

workbook.save('EAP_export_of_EDPs_401_output.xlsx')
