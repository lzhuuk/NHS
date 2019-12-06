import openpyxl


workbook1 = openpyxl.load_workbook('TemplateCheck-v1.0 - 29092019.xlsx')
worksheet1 = workbook1.worksheets[0]
workbook2 = openpyxl.load_workbook('output.xlsx')
worksheet2 = workbook2.worksheets[0]
wb = openpyxl.Workbook()
wbsheet = wb.worksheets[0]
# wb.remove_sheet('Sheet')
wbsheet['A1'] = 'ID'
wbsheet['B1'] = 'Rpt Grp Six'
wbsheet['C1'] = 'Template_previous'
wbsheet['D1'] = 'Template_now'
max_row1 = worksheet1.max_row
max_column1 = worksheet1.max_column
max_row2 = worksheet2.max_row
max_column2 = worksheet2.max_column
# data1 = []
# data2 = []
m = 1
for x in range(2, max_row1):
    cell_data1 = worksheet1.cell(row=x, column=13).value
    cell_data2 = worksheet2.cell(row=x, column=13).value
    if cell_data1 != cell_data2:
        m += 1
        wbsheet['A' + str(m)] = worksheet2.cell(row=x, column=2).value
        wbsheet['B' + str(m)] = worksheet2.cell(row=x, column=11).value
        wbsheet['C' + str(m)] = worksheet1.cell(row=x, column=12).value
        wbsheet['D' + str(m)] = worksheet2.cell(row=x, column=12).value
wb.save('compare.xlsx')
