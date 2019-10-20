import pandas as pd
import openpyxl
import os, sys

def main():

    current_path = os.path.abspath(__file__)
    os.chdir(os.path.dirname(current_path))

    df_data = pd.read_excel(\
    'results/TemplateCheck-v1.0 - 29092019.xlsx',\
    sheet_name='IN_BOTH', header=[1])

    df_data = df_data[df_data['IsSameTemplate'] == True]
    df_data = df_data[['Provider name', 'External Name', 'TemplateGene']]

    # print(df_data.head())
    # sys.exit(0)

    wb = openpyxl.load_workbook(\
    filename='sources/UCLH SER Import - Autobook (MASTER) (2)-2 copy.xlsm', \
    read_only=False, keep_vba=True)
    ws = wb['DataEntry']

    rowCount = 6
    for i, row in df_data.iterrows():
        # print(row)
        ws['H'+str(rowCount)] = row['Provider name']
        ws['J'+str(rowCount)] = row['External Name']
        ws['K'+str(rowCount)] = row['TemplateGene']
        rowCount += 1

    wb.save('results/UCLH SER Import - Autobook (MASTER) - writtenDataEntry - test.xlsm')

if __name__ == '__main__':
    print('********** Scripts start. **********')
    main()
    print('********** Scripts end. **********')
