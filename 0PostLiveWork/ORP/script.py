import openpyxl
from AddItems import addItems
from openpyxl.styles import Font
from openpyxl.styles import Alignment


fontObj1 = Font(name=u'Arial', bold=False, italic=False, size=8)
align = Alignment(vertical='top')
workbook = openpyxl.load_workbook('orp512.xlsx')
worksheet = workbook.worksheets[0]
workbook_new = openpyxl.Workbook()
worksheet_new = workbook_new.worksheets[0]
max_row = worksheet.max_row
max_col = worksheet.max_column
service = []
for x in range(6, max_row + 1):
    service.append(worksheet.cell(row=x, column=25).value)
# print(service)
row_num, rownum_urology, service_new = addItems(service)
# print(service_new)
# print(row_num)
a = 0
for i in range(0, len(row_num)):
    a += 1
    if row_num[i] not in rownum_urology:
        for j in range(1, max_col):
            worksheet_new.cell(row=a, column=j).value = worksheet.cell(row=row_num[i], column=j).value
            worksheet_new.cell(row=a, column=j).font = fontObj1
            worksheet_new.cell(row=a, column=j).alignment = align
    else:
        for j in range(1, max_col):
            worksheet_new.cell(row=a, column=j).value = worksheet.cell(row=row_num[i], column=j).value
            worksheet_new.cell(row=a, column=j).font = fontObj1
            worksheet_new.cell(row=a, column=j).alignment = align
        worksheet_new.cell(row=a, column=37).value = worksheet_new.cell(row=a, column=37).value + '\n' + '10700021'
        worksheet_new.cell(row=a, column=38).value = worksheet_new.cell(row=a, column=38).value + '\n' + \
                                                     'WMS THR EXTERNAL'
max_row0 = worksheet_new.max_row
# print(max_row0)
# print(service_new)
for y in range(1, max_row0 + 1):
    worksheet_new.cell(row=y, column=25).value = service_new[y - 1]

workbook_new.save('orp_newv3.xlsx')
