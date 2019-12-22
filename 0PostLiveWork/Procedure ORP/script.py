import re
import openpyxl
from never_appear import never_appear
from get_frequency import get_freq
from openpyxl.styles import Font
from openpyxl.styles import Alignment


fontObj1 = Font(name=u'Arial', bold=False, italic=False, size=8)
fontObj2 = Font(name=u'Arial', bold=False, italic=False, size=10)
align = Alignment(vertical='top')
workbook1 = openpyxl.load_workbook('procedure ORP.xlsx')
worksheet1 = workbook1.worksheets[0]
max_row1 = worksheet1.max_row
max_col1 = worksheet1.max_column
workbook2 = openpyxl.load_workbook('orp512.xlsx')
worksheet2 = workbook2.worksheets[0]
max_row2 = worksheet2.max_row
max_col2 = worksheet2.max_column
workbook_new = openpyxl.Workbook()
worksheet_new = workbook_new.worksheets[0]
workbook_new1 = openpyxl.Workbook()
worksheet_new1 = workbook_new1.worksheets[0]
procedure0 = []
ID = []
for x in range(2, max_row1 + 1):
    procedure0.append(worksheet1.cell(row=x, column=2).value)
for y in range(6, max_row2):
    ID.append(worksheet2.cell(row=y, column=1).value)
# print(procedure0)
procedure_ID = []
procedure = []
for i, v in enumerate(procedure0):
    procedure_ID.append(re.findall(r"\d+", str(v)))
    procedure.append(v.split(', '))
# print(procedure_ID)
# print(procedure[2][2])

# length = []
# a = 0
# num = 0
# for i in range(0, len(procedure_ID)):
#     length.append(a)
#     for j, v in enumerate(procedure_ID[i]):
#         num = len(procedure_ID[i])
#     a += num
# # print(length)
# for i, n in enumerate(length):
#     procedure.append(procedure0[i:i + 1])
# print(procedure)
# # 去除嵌套符号
# # dic = set(sum(procedure_ID, []))
# ID_NA_col = never_appear(ID, procedure_ID)
# print(ID_NA_col)
# # ID_NA_col refers to the column of procedure ID that has never been appeared
# a = 0
# for i in range(0, len(ID_NA_col)):
#     a += 1
#     for j in range(1, max_col2):
#         worksheet_new.cell(row=a, column=j).value = worksheet2.cell(row=ID_NA_col[i], column=j).value
#         worksheet_new.cell(row=a, column=j).font = fontObj1
#         worksheet_new.cell(row=a, column=j).alignment = align
# workbook_new.save('output.xlsx')

dict = get_freq(procedure)
keys, values = [], []
for key, value in sorted(dict.items(), key=lambda d: d[1], reverse=True):
    # 字典按值从大到小排列
    keys.append(key)
    values.append(value)
a = 1
worksheet_new1.cell(row=1, column=1).value = 'Procedures'
worksheet_new1.cell(row=1, column=2).value = 'Frequency'
worksheet_new1.cell(row=1, column=3).value = 'Surgeons'
worksheet_new1.cell(row=1, column=2).font = fontObj2
worksheet_new1.cell(row=1, column=2).alignment = align
worksheet_new1.cell(row=1, column=1).font = fontObj2
worksheet_new1.cell(row=1, column=1).alignment = align
worksheet_new1.cell(row=1, column=3).font = fontObj2
worksheet_new1.cell(row=1, column=3).alignment = align


for i in range(0, len(keys)):
    a += 1
    worksheet_new1.cell(row=a, column=1).value = keys[i]
    worksheet_new1.cell(row=a, column=1).font = fontObj2
    worksheet_new1.cell(row=a, column=1).alignment = align
    worksheet_new1.cell(row=a, column=2).value = values[i]
    worksheet_new1.cell(row=a, column=2).font = fontObj2
    worksheet_new1.cell(row=a, column=2).alignment = align

surgeons = []
index_num = []
for i in range(0, len(procedure)):
    name = ''
    for j, v in enumerate(procedure[i]):
        index = keys.index(v)
        # print(i + 2)
        # print(index + 2)
        if worksheet1.cell(row=i + 2, column=8).value is None:
            name += 'Unknown Person' + '\n'
            # worksheet_new1.cell(row=index + 2, column=3).value += name
        else:
            name += worksheet1.cell(row=i + 2, column=8).value + '\n'
        name.rstrip('\n')
        worksheet_new1.cell(row=index + 2, column=3).value = name
        worksheet_new1.cell(row=index + 2, column=3).font = fontObj2
        worksheet_new1.cell(row=index + 2, column=3).alignment = align
        # surgeons.append(name)
        # index_num.append(index)
        # print(worksheet1.cell(row=i + 2, column=7).value)
# print(surgeons)
# print(index_num)
workbook_new1.save('output1.xlsx')
