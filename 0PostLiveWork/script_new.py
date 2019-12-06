import openpyxl


workbook1 = openpyxl.load_workbook('ICDWHO_LEXICALS_TEXT_IMO.xlsx')
worksheet1 = workbook1.worksheets[0]
workbook2 = openpyxl.load_workbook('output(4).xlsx')
worksheet2 = workbook2.worksheets[0]
workbook3 = openpyxl.load_workbook('database.xlsx')
worksheet3 = workbook3.worksheets[0]
max_row1 = worksheet1.max_row
max_row2 = worksheet2.max_row
max_row3 = worksheet3.max_row
IMO_ID = []
SCT_ID = []
for x in range(2, max_row2 + 1):
    IMO_ID.append(str(worksheet2.cell(row=x, column=1).value))
    SCT_ID.append(str(worksheet2.cell(row=x, column=3).value))
IMO_ID1 = []
SCT_ID1 = []
for y in range(2, max_row1 + 1):
    IMO_ID1.append(str(worksheet1.cell(row=y, column=1).value))
for z in range(2, max_row3 + 1):
    SCT_ID1.append(worksheet3.cell(row=z, column=1).value)
worksheet2.cell(row=1, column=2).value = 'IMO_TEXT'
worksheet2.cell(row=1, column=4).value = 'SCT_CONCEPT_TERM'
for i in range(0, len(IMO_ID)):
    if i % 1000 == 0:
        print(i)
    if IMO_ID[i] in IMO_ID1:
        # print("In")
        index = IMO_ID1.index(IMO_ID[i])
        name = worksheet1.cell(row=index + 2, column=2).value
        worksheet2.cell(row=i + 2, column=2).value = name
    else:
        # print("Out")
        worksheet2.cell(row=i + 2, column=2).value = ''

for j in range(0, len(SCT_ID)):
    if j % 1000 == 0:
        print(j)
    if SCT_ID[j] in SCT_ID1:
        # print("In")
        index1 = SCT_ID1.index(SCT_ID[j])
        term = worksheet3.cell(row=index1 + 2, column=2).value
        worksheet2.cell(row=j + 2, column=4).value = term
    else:
        # print("Out")
        worksheet2.cell(row=j + 2, column=4).value = ''
workbook2.save('output(4).xlsx')
